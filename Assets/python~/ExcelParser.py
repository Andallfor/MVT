import openpyxl as xl
import sqlite3 as sql
import re
import sys

excel_file_name = sys.argv[1]
db_file_name = sys.argv[2]

workbook = xl.load_workbook(filename=excel_file_name, data_only=True)
conn = sql.connect(db_file_name)
cursor = conn.cursor()
first_sheet = workbook[workbook.sheetnames[0]]

for x in range(19,40,4):
    print(first_sheet.cell(1,x).value)
    name = first_sheet.cell(1,x).value+"_details"
    params = [''.join(i for i in (first_sheet.cell(2,x+y).value.replace('/','')) if not i.isdigit()) for y in range(4)]
    cursor.execute('''CREATE TABLE IF NOT EXISTS {name} ({id} PRIMARY KEY)'''.format(name=name, id=name[:-8]))
    for y in params:
        try:
            cursor.execute('''ALTER TABLE {} ADD {}'''.format(name, y))
        except:
            #print("column {} already exists".format(y))
            pass
    conn.commit()

for count, sheet in enumerate(workbook):
    #set up tables
    name = workbook.sheetnames[count].replace(" ", "_")+"_"+sheet['B1'].value.replace(" ","_")
    cursor.execute('''CREATE TABLE IF NOT EXISTS {} (Name tinytext PRIMARY KEY);'''.format(name))
    attributes = [sheet[chr(ord('A')+x)+'2'].value for x in range(0,18)]
    attributes.extend(['SBandDTE', 'SBandProximity', 'XBandDTE', 'XBandProximity', 'KaBandDTE', 'KaBandProximity', 'Schedule_Priority', 'Ground_Priority', 'Service_Level', 'Service_Period'])
    attributes.remove('Name')
    for x in attributes:
        try:
            cursor.execute('''ALTER TABLE {} ADD {}'''.format(name, x))
        except:
            #print("column {} already exists".format(x))
            pass
    conn.commit()
    #enter in data
    for row in range(3, sheet.max_row+1):
        if sheet.cell(row,1).value is None:
            break
        vals = [y.value if y.value != '\xa0' else None for y in sheet[row]]
        commandInsert = 'INSERT OR IGNORE INTO {} ('.format(name)
        commandValues = 'VALUES ('
        wasAppendedTo = False
        for col in range(1,46):
            
            if col in range(19, 43):
               if (col-19)%4 == 0:
                    wasAppendedTo = True
                    commandInsert += str(sheet.cell(1, col).value)+', '
                    commandValues += '\''+workbook.sheetnames[count].replace(' ', '_')+'_'+str(sheet.cell(row, 2).value).replace('\'','')+'\''
                    commandValues += ', '
               pass
            elif sheet.cell(row, col).value is not None and sheet.cell(row, col).value != '[]':
                wasAppendedTo = True
                commandInsert+=str(sheet.cell(2,col).value.replace(" ", "_"))+', '
                if isinstance(sheet.cell(row,col).value, str):
                    commandValues += '\''
                if col == 16:
                    m = re.search(r'([^\\]*$)', str(sheet.cell(row,col).value).replace('\'','')).group(0)
                    commandValues+=m
                else:
                    commandValues+=str(sheet.cell(row, col).value).replace('\'','')
                if isinstance(sheet.cell(row,col).value, str):
                    commandValues += '\''
                commandValues+=', '
        if wasAppendedTo:
            command = commandInsert[:-2]+")\n"+commandValues[:-2]+");"
            print(command)
            conn.execute(command)
        conn.commit()

        for tab in range(19,43, 4):
            commandInsert = 'INSERT OR IGNORE INTO {} ({}, '.format(sheet.cell(1,tab).value+"_details", sheet.cell(1,tab).value)
            commandValues = 'VALUES ({}, '.format('\''+workbook.sheetnames[count].replace(' ', '_')+'_'+str(sheet.cell(row, 2).value).replace('\'','')+'\'')
            #print(commandInsert)
            for a in range(4):
                col = tab+a
                param = ''.join(i for i in (first_sheet.cell(2, col).value.replace('/', '')) if not i.isdigit())
                if sheet.cell(row,col).value is not None and any([char.isdigit() for char in str(sheet.cell(row,col).value)]):
                    commandInsert+=param+", "
                    commandValues+=str(sheet.cell(row,col).value)+", "
                    #print("last command values: ", commandValues)
            command = commandInsert[:-2]+")\n"+commandValues[:-2]+");"
            print("Final command: ",  command)
            conn.execute(command)
        conn.commit()

conn.commit()
conn.close()



